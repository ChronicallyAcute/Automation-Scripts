import sys
import os
import re
from copy import copy

from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QListWidget, QFileDialog,
    QLabel, QProgressBar, QMessageBox,
    QRadioButton, QGroupBox
)
from PyQt5.QtCore import Qt

from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries


# --------------------------------------------------
# Formula Rewriting (Live-Link Mode)
# --------------------------------------------------

# --------------------------------------------------
# Formula Rewriting (Live-Link Mode)
# --------------------------------------------------

import re

# --- Regex components ---
CELL = r"\$?[A-Z]{1,3}\$?\d+"  # e.g., A1, $A$1, AB12, Z$99
SHEET = r"(?:'[^']+'|[A-Za-z0-9_ ]+)"  # quoted or unquoted sheet names
# Matches: Sheet!A1, 'My Sheet'!$A$1, Sheet!A1:B2, etc.
PATTERN = rf"(?P<sheet>{SHEET})!(?P<cell1>{CELL})(?::(?P<cell2>{CELL}))?"

def analyze_lock(cell_ref: str):
    """
    Analyze lock state for a single A1 ref, like: A1, $A$1, $A1, A$1.
    Returns: dict with keys col_abs, row_abs, col, row.
    """
    m = re.fullmatch(r'(\$?)([A-Z]{1,3})(\$?)(\d+)', cell_ref)
    if not m:
        return {"col_abs": False, "row_abs": False, "col": None, "row": None}
    col_abs = (m.group(1) == "$")
    col = m.group(2)
    row_abs = (m.group(3) == "$")
    row = int(m.group(4))
    return {"col_abs": col_abs, "row_abs": row_abs, "col": col, "row": row}

def strip_quotes(sheet_token: str) -> str:
    """
    Remove Excel's surrounding single quotes from a sheet token, if present,
    and unescape doubled quotes inside.
    """
    if sheet_token.startswith("'") and sheet_token.endswith("'"):
        return sheet_token[1:-1].replace("''", "'")
    return sheet_token

def quote_for_excel(name: str) -> str:
    """
    Quote and escape a sheet-or-book+sheet token for Excel.
    Always quote for safety; Excel accepts quoted even if not necessary.
    """
    return "'" + name.replace("'", "''") + "'"

def rewrite_formula_external(formula: str, source_file_name: str) -> str:
    """
    Rewrites in-workbook refs like:
        Sheet1!A1, 'My Sheet'!$A$1, Sheet1!A1:B2
    into external form:
        '[file.xlsx]Sheet1'!A1
        '[file.xlsx]My Sheet'!$A$1
        '[file.xlsx]Sheet1'!A1:B2

    - Preserves $ locks and ranges.
    - Leaves already-external refs intact.
    - Works even if a formula mixes internal+external refs (no early return).
    """

    def repl(m: re.Match) -> str:
        sheet_token = m.group("sheet")
        cell1 = m.group("cell1")
        cell2 = m.group("cell2")

        # Normalize sheet name without quotes
        sheet_name = strip_quotes(sheet_token)

        # Build the external head: [file]SheetName, then quote for Excel
        head = f"[{source_file_name}]{sheet_name}"
        head_quoted = quote_for_excel(head)

        # Preserve $ locks exactly as written; just reattach with the new head
        if cell2:
            return f"{head_quoted}!{cell1}:{cell2}"
        return f"{head_quoted}!{cell1}"

    # NOTE: We do NOT short-circuit on presence of '[' ... ']'.
    # The regex won't match already-external refs anyway (because of '['),
    # so we can safely rewrite only the internal ones in mixed formulas.
    return re.sub(PATTERN, repl, formula)


# --------------------------------------------------
# Worksheet Copy (Merged-Safe)
# --------------------------------------------------

def copy_worksheet(
    source_ws,
    target_wb,
    sheet_name,
    mode,
    source_file_name=None
):
    target_ws = target_wb.create_sheet(title=sheet_name)

    # ---- Copy merged cells first ----
    merged_ranges = list(source_ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        target_ws.merge_cells(str(merged_range))

    merged_cells = set()
    merged_anchors = set()

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        merged_anchors.add((min_row, min_col))
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                merged_cells.add((r, c))

    # ---- Copy values / formulas / styles ----
    for row in source_ws.iter_rows():
        for cell in row:
            coord = (cell.row, cell.column)

            if coord in merged_cells and coord not in merged_anchors:
                continue

            value = cell.value

            if mode == "live" and isinstance(value, str) and value.startswith("="):
                value = rewrite_formula_external(value, source_file_name)

            new_cell = target_ws.cell(
                row=cell.row,
                column=cell.column,
                value=value
            )

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # ---- Column widths ----
    for col, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col].width = dim.width

    # ---- Row heights ----
    for row, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row].height = dim.height


# --------------------------------------------------
# GUI
# --------------------------------------------------

class ExcelCompilerGUI(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Workbook Compiler")
        self.setMinimumWidth(700)

        self.selected_files = []
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        title = QLabel("Compile First Sheet of Multiple Excel Files")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-weight: bold; font-size: 15px;")
        layout.addWidget(title)

        # -------- Mode Toggle --------
        mode_box = QGroupBox("Formula Handling Mode")
        mode_layout = QVBoxLayout()

        self.static_radio = QRadioButton("Static values (no formulas, no links)")
        self.static_radio.setChecked(True)

        self.live_radio = QRadioButton("Live-linked formulas (external references)")

        mode_layout.addWidget(self.static_radio)
        mode_layout.addWidget(self.live_radio)
        mode_box.setLayout(mode_layout)

        layout.addWidget(mode_box)

        # -------- File List --------
        self.file_list = QListWidget()
        layout.addWidget(self.file_list)

        # -------- Buttons --------
        button_layout = QHBoxLayout()

        self.select_button = QPushButton("Select Excel Files")
        self.select_button.clicked.connect(self.select_files)
        button_layout.addWidget(self.select_button)

        self.compile_button = QPushButton("Compile && Save")
        self.compile_button.clicked.connect(self.compile_files)
        self.compile_button.setEnabled(False)
        button_layout.addWidget(self.compile_button)

        layout.addLayout(button_layout)

        # -------- Progress --------
        self.progress_label = QLabel("Progress:")
        layout.addWidget(self.progress_label)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        self.setLayout(layout)

    def select_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Excel Files",
            "",
            "Excel Files (*.xlsx)"
        )

        if not files:
            return

        self.selected_files = files
        self.file_list.clear()

        for f in files:
            self.file_list.addItem(os.path.basename(f))

        self.compile_button.setEnabled(True)
        self.progress_bar.setValue(0)

    def compile_files(self):
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Compiled Workbook",
            "",
            "Excel Files (*.xlsx)"
        )

        if not save_path:
            return

        mode = "static" if self.static_radio.isChecked() else "live"

        try:
            compiled_wb = Workbook()
            compiled_wb.remove(compiled_wb.active)

            self.progress_bar.setMaximum(len(self.selected_files))

            for i, file_path in enumerate(self.selected_files, start=1):
                wb = load_workbook(
                    file_path,
                    data_only=(mode == "static")
                )

                source_ws = wb.worksheets[0]
                source_file_name = os.path.basename(file_path)

                sheet_name = os.path.splitext(source_file_name)[0][:31]

                copy_worksheet(
                    source_ws,
                    compiled_wb,
                    sheet_name,
                    mode,
                    source_file_name
                )

                self.progress_bar.setValue(i)
                QApplication.processEvents()

            compiled_wb.save(save_path)

            QMessageBox.information(
                self,
                "Success",
                f"Workbook compiled successfully.\n\nMode: {mode.capitalize()}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"An error occurred:\n\n{e}"
            )


# --------------------------------------------------
# Entry Point
# --------------------------------------------------

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelCompilerGUI()
    window.show()
    sys.exit(app.exec_())
